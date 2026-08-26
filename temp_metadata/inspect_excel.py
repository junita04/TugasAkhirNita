import openpyxl

wb = openpyxl.load_workbook(r'D:\TA\TugasAkhirNita\Data\(asli)req_data_rut (baru).xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== SHEET: {sheet_name} ===')
    
    # Get headers
    rows = ws.iter_rows(max_row=1, values_only=True)
    headers = next(rows)
    print(f'Kolom ({len(headers)}): {headers}')
    
    # Count rows
    row_count = 0
    data_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_count += 1
        if i < 5:
            data_rows.append(row)
    
    print(f'Jumlah baris data: {row_count}')
    print('5 Baris pertama:')
    for r in data_rows:
        print(f'  {r}')
    
    # Check column types and nulls from first 1000 rows
    print('Analisis kolom (sample 1000 baris):')
    ws2 = wb[sheet_name]
    col_data = {h: {'null': 0, 'types': set(), 'values': []} for h in headers}
    for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True)):
        if i >= 1000:
            break
        for h, v in zip(headers, row):
            if v is None:
                col_data[h]['null'] += 1
            else:
                col_data[h]['types'].add(type(v).__name__)
                if len(col_data[h]['values']) < 5:
                    col_data[h]['values'].append(v)
    
    for h in headers:
        d = col_data[h]
        print(f'  {h}: null={d["null"]}, tipe={d["types"]}, sample={d["values"][:3]}')
    
    # Check for duplicate ID if ID column exists
    id_cols = [h for h in headers if 'id' in h.lower() or 'nim' in h.lower() or 'mhs' in h.lower()]
    if id_cols:
        ws3 = wb[sheet_name]
        seen = set()
        dup = 0
        for row in ws3.iter_rows(min_row=2, values_only=True):
            key = row[headers.index(id_cols[0])] if id_cols[0] in headers else None
            if key in seen:
                dup += 1
            seen.add(key)
        print(f'Duplicate pada {id_cols[0]}: {dup} (unique: {len(seen)})')
    
    # Check status mahasiswa unique values
    status_cols = [h for h in headers if 'status' in h.lower()]
    if status_cols:
        ws4 = wb[sheet_name]
        vals = set()
        for row in ws4.iter_rows(min_row=2, values_only=True):
            v = row[headers.index(status_cols[0])]
            if v is not None:
                vals.add(str(v).strip().upper())
        print(f'Nilai unik {status_cols[0]}: {sorted(vals)}')
    
    # Check IPK range
    ipk_cols = [h for h in headers if 'ipk' in h.lower()]
    if ipk_cols:
        ws5 = wb[sheet_name]
        vals = []
        for row in ws5.iter_rows(min_row=2, values_only=True):
            v = row[headers.index(ipk_cols[0])]
            if v is not None:
                try:
                    vals.append(float(v))
                except:
                    pass
        if vals:
            print(f'{ipk_cols[0]}: min={min(vals):.2f}, max={max(vals):.2f}, count={len(vals)}')
    
    # Check SKS range
    sks_cols = [h for h in headers if 'sks' in h.lower()]
    if sks_cols:
        ws6 = wb[sheet_name]
        vals = []
        for row in ws6.iter_rows(min_row=2, values_only=True):
            v = row[headers.index(sks_cols[0])]
            if v is not None:
                try:
                    vals.append(int(v))
                except:
                    pass
        if vals:
            print(f'{sks_cols[0]}: min={min(vals)}, max={max(vals)}, count={len(vals)}')
    
    # Check MK range
    mk_cols = [h for h in headers if 'mk' in h.lower() or 'matkul' in h.lower() or 'mata kuliah' in h.lower()]
    if mk_cols:
        ws7 = wb[sheet_name]
        vals = []
        for row in ws7.iter_rows(min_row=2, values_only=True):
            v = row[headers.index(mk_cols[0])]
            if v is not None:
                try:
                    vals.append(int(v))
                except:
                    pass
        if vals:
            print(f'{mk_cols[0]}: min={min(vals)}, max={max(vals)}, count={len(vals)}')

wb.close()